"""BEA I-O XLSX loader for tensor hierarchy feature.

Feature: 025-tensor-hierarchy
Date: 2026-02-26

Parses BEA Input-Output Use tables (IOUse_Before_Redefinitions_PRO_Summary.xlsx)
and loads direct requirements coefficients into the 3NF schema.

Structure of BEA I-O XLSX files:
    - Multiple sheets: one per year (e.g. '1997', '1998', ..., '2024')
    - Row 1: Title string
    - Row 2: Units "(Millions of dollars)"
    - Row 3: "Bureau of Economic Analysis"
    - Row 4: Year string
    - Row 5: Empty
    - Row 6: None, "Commodities/Industries", <BEA codes...>
    - Row 7: "IOCode", "Name", <industry names...>
    - Rows 8+: <IOCode>, <Name>, <values or '...'>
    - Special rows: T001 (total intermediate), V001 (value added), T019 (gross output)

Missing data is marked as string '...' and treated as 0.

Usage:
    from babylon.data.bea.io_loader import BEAIOLoader
    loader = BEAIOLoader()
    stats = loader.load(session)

See Also:
    :mod:`babylon.data.loader_base`: DataLoader base class.
    :mod:`babylon.data.reference.schema`: DimBEAIOTableType, FactBEAIOCoefficient.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np

from babylon.data.loader_base import DataLoader, LoaderConfig, LoadStats
from babylon.data.reference.schema import (
    DimBEAIndustry,
    DimBEAIOTableType,
    FactBEAIOCoefficient,
)

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# Row indices (0-based) in the BEA I-O XLSX format
_ROW_TITLE = 0
_ROW_UNITS = 1
_ROW_SOURCE = 2
_ROW_YEAR = 3
_ROW_EMPTY = 4
_ROW_CODES = 5  # column codes start at index 2
_ROW_NAMES = 6  # column names
_ROW_DATA_START = 7

# Row IOCode prefixes that represent summary/total rows (not industry rows)
_SKIP_CODES = frozenset(
    [
        "T001",
        "T002",
        "T003",
        "T004",
        "T005",
        "T006",
        "T007",
        "T008",
        "T009",
        "T010",
        "T011",
        "T012",
        "T013",
        "T014",
        "T015",
        "T016",
        "T017",
        "T018",
        "T019",
        "T020",
        "V001",
        "V002",
        "V003",
        "V004",
        "V005",
        "V006",
    ]
)

# Default XLSX filename (Summary Use table, Producers' prices)
_XLSX_FILENAME = "IOUse_Before_Redefinitions_PRO_Summary.xlsx"

# Default data directory relative to project root
DEFAULT_DATA_DIR = Path("data")


class IOMatrixParser:
    """Parser for BEA I-O coefficient matrices from XLSX row data.

    Accepts rows as tuples (values_only=True from openpyxl) and extracts:
    - Ordered list of BEA industry codes (columns)
    - Direct requirements coefficient matrix (normalized by gross output)

    Example:
        >>> parser = IOMatrixParser()
        >>> rows = [...]  # from openpyxl ws.iter_rows(values_only=True)
        >>> industries, matrix = parser.parse_rows(rows)
    """

    def parse_rows(self, rows: list[tuple[object, ...]]) -> tuple[list[str], np.ndarray]:
        """Parse BEA I-O XLSX rows into industry list and coefficient matrix.

        The Use table contains the dollar value of commodity i used by industry j.
        Direct requirements coefficient A[i,j] = Use[i,j] / GrossOutput[j].

        Args:
            rows: List of row tuples from openpyxl iter_rows(values_only=True).
                  Expected format: title rows 0-5, data rows 6+.

        Returns:
            Tuple of:
            - industries: Ordered list of BEA industry codes (columns).
            - matrix: Direct requirements matrix A, shape (n, n), float64.

        Note:
            Missing data '...' is converted to 0.0.
            Total/summary rows (T0xx, V0xx) are excluded from industry list.
        """
        if len(rows) < _ROW_DATA_START + 1:
            return [], np.zeros((0, 0), dtype=np.float64)

        col_codes = self._extract_col_codes(rows)
        if not col_codes:
            return [], np.zeros((0, 0), dtype=np.float64)

        row_codes, raw_values, gross_output = self._parse_data_section(rows, col_codes)
        if not row_codes:
            return [], np.zeros((0, 0), dtype=np.float64)

        a_matrix = self._build_coeff_matrix(row_codes, raw_values, col_codes, gross_output)
        return col_codes, a_matrix

    def _extract_col_codes(self, rows: list[tuple[object, ...]]) -> list[str]:
        """Extract industry column codes from header row, excluding final-demand F-codes.

        Args:
            rows: Full XLSX row list (0-indexed).

        Returns:
            Ordered list of BEA industry codes (F-prefixed final-demand codes excluded).
        """
        code_row = rows[_ROW_CODES]
        return [
            str(v)
            for v in code_row[2:]
            if v is not None
            and str(v).strip() not in ("", "None")
            and not str(v).strip().startswith("F")  # exclude final demand columns
        ]

    def _parse_data_section(
        self,
        rows: list[tuple[object, ...]],
        col_codes: list[str],
    ) -> tuple[list[str], list[list[float]], dict[str, float]]:
        """Scan data rows to collect industry flows and derive gross output.

        Uses T019 if present; otherwise accumulates commodity + value-added rows
        as a column-sum proxy (GrossOutput = intermediate use + value added).

        Args:
            rows: Full XLSX row list.
            col_codes: Ordered list of industry column codes.

        Returns:
            Tuple of (row_codes, raw_values, gross_output) where gross_output
            maps industry code to estimated gross output (always >= 1.0).
        """
        n_cols = len(col_codes)
        row_codes: list[str] = []
        raw_values: list[list[float]] = []
        gross_output: dict[str, float] = {}
        col_accum: list[float] = [0.0] * n_cols

        for row in rows[_ROW_DATA_START:]:
            if not row or row[0] is None:
                continue
            io_code = str(row[0]).strip()
            if not io_code or io_code == "None":
                continue

            vals = [self._to_float(row[i + 2]) for i in range(n_cols)]

            if io_code == "T019":
                gross_output = {col_codes[i]: max(vals[i], 1.0) for i in range(n_cols)}
                continue

            if io_code in _SKIP_CODES:
                if io_code.startswith("V"):
                    for i in range(n_cols):
                        col_accum[i] += vals[i]
                continue

            row_codes.append(io_code)
            raw_values.append(vals)
            for i in range(n_cols):
                col_accum[i] += vals[i]

        if not gross_output:
            gross_output = {col_codes[i]: max(col_accum[i], 1.0) for i in range(n_cols)}

        return row_codes, raw_values, gross_output

    def _build_coeff_matrix(
        self,
        row_codes: list[str],
        raw_values: list[list[float]],
        col_codes: list[str],
        gross_output: dict[str, float],
    ) -> np.ndarray:
        """Build and normalize the n×n direct-requirements A matrix.

        Args:
            row_codes: IOCodes for each data row (may include non-industry codes).
            raw_values: Parallel list of float value lists.
            col_codes: Canonical industry column codes (defines matrix dimensions).
            gross_output: Maps each industry code to its gross output value.

        Returns:
            Direct requirements matrix A, shape (n, n), float64,
            normalized by column gross output.
        """
        n = len(col_codes)
        code_to_col: dict[str, int] = {c: i for i, c in enumerate(col_codes)}
        a_matrix = np.zeros((n, n), dtype=np.float64)

        for r_idx, code in enumerate(row_codes):
            row_pos = code_to_col.get(code)
            if row_pos is None:
                continue
            vals = raw_values[r_idx]
            for c_idx in range(min(n, len(vals))):
                a_matrix[row_pos, c_idx] = vals[c_idx]

        for c_idx, code in enumerate(col_codes):
            output = gross_output.get(code, 0.0)
            if output > 0.0:
                a_matrix[:, c_idx] /= output

        return a_matrix

    @staticmethod
    def _to_float(value: object) -> float:
        """Convert cell value to float, treating '...' (missing) as 0.0.

        Args:
            value: Cell value from openpyxl (int, float, str, or None).

        Returns:
            float value, 0.0 for missing data markers.
        """
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if s in ("...", "", "None", "-"):
            return 0.0
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return 0.0


class BEAIOLoader(DataLoader):
    """Loader for BEA I-O coefficients into 3NF schema.

    Reads IOUse_Before_Redefinitions_PRO_Summary.xlsx (multi-sheet XLSX,
    one sheet per year 1997-2024) and populates:
    - dim_bea_io_table_type: USE table type record
    - fact_bea_io_coefficient: Direct requirements A[i,j] by year

    Requires dim_bea_industry to be pre-populated (by BEANationalLoader).
    Uses existing dim_time records or creates them as needed.

    Args:
        config: LoaderConfig for operational settings.
        data_dir: Path to data directory containing input-output/make-use/.

    Example:
        loader = BEAIOLoader()
        with session_factory() as session:
            stats = loader.load(session)
    """

    def __init__(
        self,
        config: LoaderConfig | None = None,
        data_dir: Path | None = None,
    ) -> None:
        """Initialize BEA I-O loader.

        Args:
            config: LoaderConfig for operational settings.
            data_dir: Base data directory. Defaults to "data" in project root.
        """
        super().__init__(config)
        self.data_dir = data_dir if data_dir is not None else DEFAULT_DATA_DIR
        self._parser = IOMatrixParser()

    def get_dimension_tables(self) -> list[type]:
        """Return dimension tables this loader populates.

        Returns:
            List with DimBEAIOTableType.
        """
        return [DimBEAIOTableType]

    def get_fact_tables(self) -> list[type]:
        """Return fact tables this loader populates.

        Returns:
            List with FactBEAIOCoefficient.
        """
        return [FactBEAIOCoefficient]

    def load(
        self,
        session: Session,
        reset: bool = True,
        verbose: bool = True,
        **kwargs: object,  # noqa: ARG002
    ) -> LoadStats:
        """Load BEA I-O coefficients into 3NF schema.

        Parses the Summary-level Use table XLSX and inserts/replaces direct
        requirements coefficients by (year, source_industry, target_industry).

        Args:
            session: SQLAlchemy session for the normalized database.
            reset: If True, delete existing records before loading.
            verbose: If True, print progress information.
            **kwargs: Ignored.

        Returns:
            LoadStats with counts of loaded dimensions and facts.
        """
        stats = LoadStats(source="bea_io")
        xlsx_path = self.data_dir / "input-output" / "make-use" / _XLSX_FILENAME

        if not xlsx_path.exists():
            msg = f"BEA I-O XLSX not found: {xlsx_path}"
            logger.warning(msg)
            stats.errors.append(msg)
            return stats

        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError as exc:
            msg = f"openpyxl required for BEA I-O loading: {exc}"
            stats.errors.append(msg)
            return stats

        with session.begin():
            if reset:
                self.clear_tables(session)

            # Ensure USE table type exists
            table_type_id = self._get_or_create_table_type(session)
            stats.dimensions_loaded["dim_bea_io_table_type"] = 1

            # Build industry code -> bea_industry_id lookup
            industry_lookup = self._build_industry_lookup(session)
            if not industry_lookup:
                msg = "dim_bea_industry is empty; run BEANationalLoader first"
                logger.warning(msg)
                stats.errors.append(msg)
                return stats

            # Parse each year sheet
            try:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            except Exception as exc:
                msg = f"Failed to open {xlsx_path}: {exc}"
                stats.errors.append(msg)
                return stats

            total_coefficients = 0
            try:
                for sheet_name in wb.sheetnames:
                    year = self._parse_year(sheet_name)
                    if year is None:
                        continue

                    ws = wb[sheet_name]
                    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
                    industries, matrix = self._parser.parse_rows(rows)

                    if not industries:
                        logger.debug("No industries parsed for sheet %s", sheet_name)
                        continue

                    time_id = self._get_or_create_time(session, year)
                    count = self._load_coefficients(
                        session,
                        year=year,
                        time_id=time_id,
                        table_type_id=table_type_id,
                        industries=industries,
                        matrix=matrix,
                        industry_lookup=industry_lookup,
                    )
                    total_coefficients += count
                    stats.files_processed += 1
                    if verbose:
                        logger.debug("Loaded %d coefficients for %d", count, year)
            finally:
                wb.close()

        stats.facts_loaded["fact_bea_io_coefficient"] = total_coefficients
        return stats

    def _get_or_create_table_type(self, session: Session) -> int:
        """Get or create the USE table type record.

        Args:
            session: SQLAlchemy session.

        Returns:
            id of the DimBEAIOTableType record.
        """
        existing = (
            session.query(DimBEAIOTableType).filter(DimBEAIOTableType.table_type == "USE").first()
        )
        if existing:
            return existing.id

        record = DimBEAIOTableType(
            table_type="USE",
            description="Use table (Producers' prices): commodity use by industry",
        )
        session.add(record)
        session.flush()
        return record.id

    def _build_industry_lookup(self, session: Session) -> dict[str, int]:
        """Build BEA code -> bea_industry_id lookup from dim_bea_industry.

        Args:
            session: SQLAlchemy session.

        Returns:
            Dict mapping bea_code -> bea_industry_id.
        """
        industries = session.query(DimBEAIndustry).all()
        return {ind.bea_code: ind.bea_industry_id for ind in industries}

    def _load_coefficients(
        self,
        session: Session,
        year: int,  # noqa: ARG002
        time_id: int,
        table_type_id: int,
        industries: list[str],
        matrix: np.ndarray,
        industry_lookup: dict[str, int],
    ) -> int:
        """Insert direct requirements coefficients for one year.

        Args:
            session: SQLAlchemy session.
            year: Data year.
            time_id: FK to dim_time.
            table_type_id: FK to dim_bea_io_table_type.
            industries: Ordered list of BEA codes (columns/rows).
            matrix: Direct requirements matrix A, shape (n, n).
            industry_lookup: BEA code -> bea_industry_id.

        Returns:
            Number of coefficient rows inserted.
        """
        n = len(industries)
        records: list[dict[str, object]] = []

        for src_idx in range(n):
            src_code = industries[src_idx]
            src_id = industry_lookup.get(src_code)
            if src_id is None:
                continue

            for tgt_idx in range(n):
                tgt_code = industries[tgt_idx]
                tgt_id = industry_lookup.get(tgt_code)
                if tgt_id is None:
                    continue

                coeff = float(matrix[src_idx, tgt_idx])
                if coeff == 0.0:
                    continue  # skip zero entries for sparse storage

                records.append(
                    {
                        "time_id": time_id,
                        "table_type_id": table_type_id,
                        "source_industry_id": src_id,
                        "target_industry_id": tgt_id,
                        "coefficient": coeff,
                    }
                )

        if not records:
            return 0

        return self._upsert_batch(
            session,
            FactBEAIOCoefficient,
            records,
            index_elements=["time_id", "table_type_id", "source_industry_id", "target_industry_id"],
            update_columns=["coefficient"],
            batch_size=self.config.batch_size,
        )

    @staticmethod
    def _parse_year(sheet_name: str) -> int | None:
        """Parse sheet name to year integer.

        Args:
            sheet_name: XLSX sheet name (e.g. '1997', '2021').

        Returns:
            Year as int, or None if not a valid year string.
        """
        try:
            year = int(sheet_name.strip())
            if 1990 <= year <= 2030:
                return year
        except ValueError:
            pass
        return None


__all__ = [
    "BEAIOLoader",
    "IOMatrixParser",
]
