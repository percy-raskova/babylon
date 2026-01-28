"""BEA XLSX file parser for GDP-by-industry data.

Parses the official BEA XLSX workbooks containing:
- GrossOutput.xlsx (TGO105-A sheet)
- ValueAdded.xlsx (TVA105-A sheet)
- IntermediateInputs.xlsx (TII105-A sheet)

These files share a common structure:
- Rows 0-6: Header metadata
- Row 7: Column headers (Line, years...)
- Row 8+: Industry data with hierarchical indentation

The parser extracts:
- Industry hierarchy (via indentation level in industry names)
- Annual dollar values by industry and year
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook  # type: ignore[import-untyped]
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

logger = logging.getLogger(__name__)

# BEA data file locations (relative to data directory)
BEA_DATA_DIR = "gdp-by-industry"
GROSS_OUTPUT_FILE = "GrossOutput.xlsx"
VALUE_ADDED_FILE = "ValueAdded.xlsx"
INTERMEDIATE_INPUTS_FILE = "IntermediateInputs.xlsx"

# Sheet names containing annual dollar values
GROSS_OUTPUT_SHEET = "TGO105-A"
VALUE_ADDED_SHEET = "TVA105-A"
INTERMEDIATE_INPUTS_SHEET = "TII105-A"

# Row where data headers start (0-indexed)
HEADER_ROW = 7
DATA_START_ROW = 8


@dataclass
class BEAIndustry:
    """Parsed BEA industry with hierarchy information.

    Attributes:
        line_number: BEA table line number (unique identifier).
        name: Industry name (with indentation stripped).
        raw_name: Original industry name with indentation.
        level: Hierarchy level (1=sector, 2=major, 3=industry, 4=detail).
        code: Generated code from line number (e.g., "BEA001").
    """

    line_number: int
    name: str
    raw_name: str
    level: int
    code: str


@dataclass
class BEAIndustryValue:
    """Annual value for a BEA industry.

    Attributes:
        line_number: BEA table line number (links to BEAIndustry).
        year: Calendar year.
        value_millions: Dollar value in millions.
    """

    line_number: int
    year: int
    value_millions: Decimal | None


@dataclass
class BEAParseResult:
    """Complete parse result from a BEA XLSX file.

    Attributes:
        industries: List of parsed industries with hierarchy.
        values: List of annual values by industry.
        years: List of years present in the data.
        source_file: Name of the source file.
        value_type: Type of value (gross_output, value_added, intermediate_inputs).
    """

    industries: list[BEAIndustry] = field(default_factory=list)
    values: list[BEAIndustryValue] = field(default_factory=list)
    years: list[int] = field(default_factory=list)
    source_file: str = ""
    value_type: str = ""


class BEAIndustryParser:
    """Parser for BEA GDP-by-industry XLSX files.

    Extracts industry hierarchy and annual values from the official BEA
    workbooks. Handles the indentation-based hierarchy in industry names
    and maps to a consistent level scheme.

    Example:
        parser = BEAIndustryParser(Path("data"))
        result = parser.parse_gross_output()
        for industry in result.industries:
            print(f"{industry.code}: {industry.name} (level {industry.level})")
    """

    def __init__(self, data_dir: Path) -> None:
        """Initialize parser with data directory.

        Args:
            data_dir: Path to the data directory containing gdp-by-industry/.
        """
        self.data_dir = data_dir
        self.bea_dir = data_dir / BEA_DATA_DIR

    def parse_gross_output(self) -> BEAParseResult:
        """Parse GrossOutput.xlsx and return industries with values."""
        filepath = self.bea_dir / GROSS_OUTPUT_FILE
        return self._parse_workbook(filepath, GROSS_OUTPUT_SHEET, "gross_output")

    def parse_value_added(self) -> BEAParseResult:
        """Parse ValueAdded.xlsx and return industries with values."""
        filepath = self.bea_dir / VALUE_ADDED_FILE
        return self._parse_workbook(filepath, VALUE_ADDED_SHEET, "value_added")

    def parse_intermediate_inputs(self) -> BEAParseResult:
        """Parse IntermediateInputs.xlsx and return industries with values."""
        filepath = self.bea_dir / INTERMEDIATE_INPUTS_FILE
        return self._parse_workbook(filepath, INTERMEDIATE_INPUTS_SHEET, "intermediate_inputs")

    def _parse_workbook(self, filepath: Path, sheet_name: str, value_type: str) -> BEAParseResult:
        """Parse a BEA workbook and extract industries and values.

        Args:
            filepath: Path to the XLSX file.
            sheet_name: Name of the sheet containing annual data.
            value_type: Type identifier for the values.

        Returns:
            BEAParseResult with industries and values.
        """
        import openpyxl  # type: ignore[import-untyped]

        if not filepath.exists():
            raise FileNotFoundError(f"BEA data file not found: {filepath}")

        logger.info(f"Parsing {filepath.name} sheet {sheet_name}")

        wb: Workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet {sheet_name} not found in {filepath.name}")

        ws: Worksheet = wb[sheet_name]
        result = BEAParseResult(
            source_file=filepath.name,
            value_type=value_type,
        )

        # Extract all rows as values
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) <= DATA_START_ROW:
            logger.warning(f"No data rows found in {filepath.name}")
            return result

        # Parse header row to get years
        header_row = rows[HEADER_ROW]
        result.years = self._parse_years(header_row)
        logger.info(f"Found years: {result.years[0]} - {result.years[-1]}")

        # Parse data rows
        seen_industries: set[int] = set()
        for row in rows[DATA_START_ROW:]:
            if not row or row[0] is None:
                continue

            # First column is line number, second is industry name
            try:
                line_num = int(row[0])
            except (ValueError, TypeError):
                continue

            # Get industry name from column 1 (may be column 2 in some sheets)
            industry_name = self._get_industry_name(row)
            if not industry_name:
                continue

            # Parse industry if not seen
            if line_num not in seen_industries:
                industry = self._parse_industry(line_num, industry_name)
                result.industries.append(industry)
                seen_industries.add(line_num)

            # Parse values for each year
            year_values = self._extract_year_values(row, result.years)
            for year, value in year_values.items():
                result.values.append(
                    BEAIndustryValue(
                        line_number=line_num,
                        year=year,
                        value_millions=value,
                    )
                )

        logger.info(
            f"Parsed {len(result.industries)} industries, "
            f"{len(result.values)} values from {filepath.name}"
        )
        return result

    def _parse_years(self, header_row: tuple[object, ...]) -> list[int]:
        """Extract years from header row."""
        years: list[int] = []
        for cell in header_row:
            if cell is None:
                continue
            try:
                # Convert to string first to satisfy mypy, then to int
                year = int(str(cell))
                if 1990 <= year <= 2050:  # Reasonable year range
                    years.append(year)
            except (ValueError, TypeError):
                continue
        return sorted(years)

    def _get_industry_name(self, row: tuple[object, ...]) -> str | None:
        """Extract industry name from row (handles different column layouts).

        Note: Preserves leading spaces for hierarchy detection.
        """
        # Usually column 1 (index 1) contains the industry name
        if len(row) > 1 and row[1] is not None:
            name = str(row[1])  # Don't strip - preserve leading spaces
            if name.strip():  # But check if there's actual content
                return name
        return None

    def _parse_industry(self, line_number: int, raw_name: str) -> BEAIndustry:
        """Parse industry with hierarchy level from indentation.

        BEA indentation scheme:
        - Line 1 "All industries" (4 spaces): level 1 (total economy)
        - 0 spaces (e.g., "Private industries"): level 1 (major categories)
        - 2 spaces (e.g., "  Agriculture..."): level 2 (major sectors)
        - 4 spaces (e.g., "    Farms"): level 3 (industries)
        - 6 spaces (e.g., "      Wood products"): level 4 (detail industries)
        """
        # Count leading spaces to determine level
        stripped = raw_name.lstrip()
        indent = len(raw_name) - len(stripped)

        # Map indentation to hierarchy level
        # Line 1 is special case - always level 1 (total economy)
        if line_number == 1:
            level = 1
        elif indent == 0:
            level = 1  # Major categories like "Private industries", "Government"
        elif indent == 2:
            level = 2  # Sectors like "Agriculture", "Mining"
        elif indent == 4:
            level = 3  # Industries like "Farms", "Oil and gas extraction"
        else:
            level = 4  # Detail industries (6+ spaces)

        # Generate code from line number
        code = f"BEA{line_number:03d}"

        return BEAIndustry(
            line_number=line_number,
            name=stripped,
            raw_name=raw_name,
            level=level,
            code=code,
        )

    def _extract_year_values(
        self, row: tuple[object, ...], years: list[int]
    ) -> dict[int, Decimal | None]:
        """Extract values for each year from a data row.

        Args:
            row: Row tuple from the worksheet.
            years: List of years (in order they appear in columns).

        Returns:
            Dictionary mapping year to value (or None if missing).
        """
        result: dict[int, Decimal | None] = {}

        # Values start after line number and industry name columns
        # In TGO105-A format: Line, IndustryName, (empty), Year1, Year2, ...
        # So values start at column index 3 (0-indexed)
        value_start_col = 2  # Adjusted based on actual data structure

        # Find where the numeric year values actually start
        for idx, cell in enumerate(row):
            if idx < 2:  # Skip line number and industry name
                continue
            if cell is not None and isinstance(cell, (int, float)):
                value_start_col = idx
                break

        # Now extract values aligned with years
        for i, year in enumerate(years):
            col_idx = value_start_col + i
            if col_idx < len(row):
                cell = row[col_idx]
                if cell is not None:
                    try:
                        value = Decimal(str(cell))
                        result[year] = value
                    except Exception:
                        result[year] = None
                else:
                    result[year] = None
            else:
                result[year] = None

        return result


__all__ = [
    "BEAIndustryParser",
    "BEAIndustry",
    "BEAIndustryValue",
    "BEAParseResult",
]
