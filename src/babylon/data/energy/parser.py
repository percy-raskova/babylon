"""EIA Monthly Energy Review Excel parser.

Parses annual energy data from EIA MER Excel files into structured records
for database loading.

Excel structure (observed from Table 01.01.xlsx):
    Row 0: "U.S. Energy Information Administration"
    Row 1: "December 2025 Monthly Energy Review"
    Row 3-4: Release/update dates
    Row 6: Table title (e.g., "Table 1.1 Primary energy overview")
    Row 8: Column headers (first column is "Annual Total" = year)
    Row 9: Units in parentheses (e.g., "(Quadrillion Btu)")
    Row 10: Blank
    Rows 11+: Data (year in first column, values in subsequent columns)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl  # type: ignore[import-untyped]

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]


@dataclass
class EnergyRecord:
    """Parsed annual energy observation."""

    table_code: str
    series_name: str
    units: str
    column_index: int
    year: int
    value: float | None


@dataclass
class EnergyTableData:
    """Parsed data from a single EIA table."""

    table_code: str
    table_title: str
    series: list[dict[str, str | int]] = field(default_factory=list)
    records: list[EnergyRecord] = field(default_factory=list)


def extract_table_code(filename: str) -> str:
    """Extract table code from filename.

    Args:
        filename: Excel filename (e.g., "Table 01.01.xlsx")

    Returns:
        Table code (e.g., "01.01")

    Examples:
        >>> extract_table_code("Table 01.01.xlsx")
        '01.01'
        >>> extract_table_code("Table 09.04.xlsx")
        '09.04'
        >>> extract_table_code("Table 02.01a.xlsx")
        '02.01a'
    """
    match = re.search(r"Table\s+(\d+\.\d+[a-z]?)", filename, re.IGNORECASE)
    if match:
        return match.group(1)
    # Fallback: use filename without extension
    return Path(filename).stem.replace("Table ", "").strip()


def parse_units(raw_units: str | None) -> str:
    """Parse units string, removing parentheses.

    Args:
        raw_units: Raw units string (e.g., "(Quadrillion Btu)")

    Returns:
        Cleaned units string (e.g., "Quadrillion Btu")

    Examples:
        >>> parse_units("(Quadrillion Btu)")
        'Quadrillion Btu'
        >>> parse_units("Dollars per Barrel")
        'Dollars per Barrel'
        >>> parse_units(None)
        'Unknown'
    """
    if not raw_units:
        return "Unknown"
    # Remove parentheses
    cleaned = str(raw_units).strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = cleaned[1:-1]
    return cleaned.strip() or "Unknown"


def parse_value(raw_value: str | float | int | None) -> float | None:
    """Parse cell value to float, handling special cases.

    EIA uses various non-numeric markers:
        - "Not Available" or "NA" -> None
        - Empty or None -> None
        - Numeric values -> float

    Args:
        raw_value: Raw cell value

    Returns:
        Float value or None

    Examples:
        >>> parse_value(28.740479)
        28.740479
        >>> parse_value("Not Available")
        >>> parse_value(None)
        >>> parse_value("")
    """
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    raw_str = str(raw_value).strip()
    if not raw_str or raw_str.lower() in ("not available", "na", "n/a", "--", "-"):
        return None
    try:
        return float(raw_str.replace(",", ""))
    except ValueError:
        return None


def parse_energy_excel(
    file_path: Path,
    sheet_name: str = "Annual Data",
    header_row: int = 8,
    units_row: int = 9,
    data_start_row: int = 11,
) -> EnergyTableData:
    """Parse EIA MER Excel file into structured data.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet to read (default: "Annual Data")
        header_row: Row index for column headers (0-indexed)
        units_row: Row index for units (0-indexed)
        data_start_row: Row index where data begins (0-indexed)

    Returns:
        EnergyTableData containing series metadata and records

    Raises:
        FileNotFoundError: If file does not exist
        KeyError: If sheet_name not found in workbook
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Energy file not found: {file_path}")

    table_code = extract_table_code(file_path.name)

    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Sheet '{sheet_name}' not found in {file_path.name}")

    ws: Worksheet = wb[sheet_name]

    # Extract table title from row 6
    table_title = str(ws.cell(row=7, column=1).value or f"Table {table_code}")

    # Extract headers and units
    headers: list[str | None] = []
    units: list[str] = []

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row + 1, column=col).value
        units_val = ws.cell(row=units_row + 1, column=col).value
        headers.append(str(header_val) if header_val else None)
        units.append(parse_units(units_val))

    # Build series metadata (skip first column which is year)
    series_list: list[dict[str, str | int]] = []
    for col_idx in range(1, len(headers)):
        header = headers[col_idx]
        if header and header.strip():
            series_list.append(
                {
                    "series_name": header.strip(),
                    "units": units[col_idx] if col_idx < len(units) else "Unknown",
                    "column_index": col_idx,
                }
            )

    # Parse data rows
    records: list[EnergyRecord] = []
    for row in ws.iter_rows(min_row=data_start_row + 1, values_only=True):
        if not row or row[0] is None:
            continue

        # First column is year
        try:
            year = int(row[0])
        except (ValueError, TypeError):
            continue

        # Parse each series column
        for series_info in series_list:
            col_idx = int(series_info["column_index"])
            if col_idx < len(row):
                value = parse_value(row[col_idx])
                records.append(
                    EnergyRecord(
                        table_code=table_code,
                        series_name=str(series_info["series_name"]),
                        units=str(series_info["units"]),
                        column_index=int(col_idx),
                        year=year,
                        value=value,
                    )
                )

    wb.close()

    return EnergyTableData(
        table_code=table_code,
        table_title=table_title,
        series=series_list,
        records=records,
    )


def discover_energy_files(
    energy_dir: Path,
    priority_tables: set[str] | None = None,
) -> dict[str, Path]:
    """Discover EIA Excel files in directory.

    Args:
        energy_dir: Path to data/energy/ directory
        priority_tables: Set of table codes to filter (e.g., {"01.01", "09.01"}).
            If None, discovers all Table *.xlsx files.

    Returns:
        Dict mapping table_code to file path
    """
    if not energy_dir.exists():
        return {}

    files: dict[str, Path] = {}
    for xlsx_path in energy_dir.glob("Table *.xlsx"):
        table_code = extract_table_code(xlsx_path.name)
        if priority_tables is None or table_code in priority_tables:
            files[table_code] = xlsx_path

    return files
