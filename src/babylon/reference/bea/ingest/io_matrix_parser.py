"""Parser for the BEA Make+Use IOUse_Before_Redefinitions_PRO_Summary XLSX.

Spec-068 US2. File path:
``data/input-output/make-use/IOUse_Before_Redefinitions_PRO_Summary.xlsx``.

Sheet structure (per-year, identical layout to Supply-Use Use_Summary):

* R1: title; R2: ``"(Millions of dollars)"``; R3: BEA; R4: year.
* R6: BEA codes in cols 3+ (column = target industry j).
* R7: industry names in cols 3+.
* Rows 8+: per-input-row dollar values. Col 0 = IOCode (input industry i),
  Col 1 = name, Cols 2+ = dollar use of i by industry j.

Special stub rows (per Use-table convention):

* ``T005``: Total Intermediate Use (row sum).
* ``T018``: Total industry output at basic prices — the GO denominator.

To convert raw USE dollar values to direct-requirements coefficients
``a_ij``:

    a_ij = USE_dollars[i, j] / T018[j]

That gives "dollars of input industry i required per dollar of output
industry j" — the Leontief direct-requirements form (FR-003).

Sparsity policy (per data-model.md §Entity 2): cells where
``USE_dollars[i,j] == 0`` or ``'...'`` are omitted — they are not
written as zero rows.
"""

from __future__ import annotations

import logging
from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.orm import Session

from babylon.reference.bea.ingest.supply_use_parser import (
    BEAIngestError,
    _coerce_cell_to_decimal,
)
from babylon.reference.bea.ingest.vintage_extractor import extract_vintage_date
from babylon.reference.bea.models import BEAIOCoefficientRecord
from babylon.reference.schema import DimBEAIndustry, DimTime

log = logging.getLogger(__name__)

_BEA_CODE_ROW = 6
_FIRST_INDUSTRY_COL = 3
_IOCODE_COL = 1
_NAME_COL = 2
_STUB_GROSS_OUTPUT_IOCODE = "T018"
_STUB_GROSS_OUTPUT_NAME = "Total Industry Output"  # IOUse uses Name when IOCode is None


def parse_use_matrix(
    xlsx_path: Path,
    years: range,
    session: Session,
    table_type: str = "USE",
) -> Iterator[BEAIOCoefficientRecord]:
    """Yield ``BEAIOCoefficientRecord`` per non-zero (i, j) cell.

    Args:
        xlsx_path: Path to ``IOUse_Before_Redefinitions_PRO_Summary.xlsx``.
        years: Year range to ingest.
        session: SQLAlchemy session for ``dim_bea_industry`` + ``dim_time`` lookups.
        table_type: ``"USE"`` (default) — discriminator written to the
            ``dim_bea_io_table_type`` join. Other values are accepted
            for future Make/Supply/Import-Use ingest.

    Yields:
        Frozen Pydantic records ready for the UPSERT writer.

    Raises:
        BEAIngestError: If the XLSX is missing or the T018 stub row is absent.
    """
    if not xlsx_path.exists():
        raise BEAIngestError(f"BEA IOUse XLSX not found at {xlsx_path!r}")

    vintage = extract_vintage_date(xlsx_path)
    bea_rows = session.execute(
        select(DimBEAIndustry.bea_code, DimBEAIndustry.bea_industry_id)
    ).all()
    bea_code_lookup: dict[str, int] = {code: bid for code, bid in bea_rows}  # noqa: C416
    year_rows = session.execute(
        select(DimTime.year, DimTime.time_id).where(DimTime.is_annual.is_(True))
    ).all()
    year_lookup: dict[int, int] = {year: tid for year, tid in year_rows}  # noqa: C416
    years_in_scope = set(years)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                continue
            if year not in years_in_scope or year not in year_lookup:
                continue
            ws = wb[sheet_name]
            yield from _parse_year_matrix(
                ws=ws,
                year=year,
                bea_code_lookup=bea_code_lookup,
                vintage=vintage,
                table_type=table_type,
            )
    finally:
        wb.close()


def _parse_year_matrix(  # noqa: C901 — XLSX multi-pass parser
    ws: object,
    year: int,
    bea_code_lookup: dict[str, int],
    vintage: object,
    table_type: str,
) -> Iterator[BEAIOCoefficientRecord]:
    """Yield coefficient records for one year sheet."""
    # First pass: industry column → BEA code
    industry_col_to_bea_code: dict[int, str] = {}
    code_row = next(ws.iter_rows(min_row=_BEA_CODE_ROW, max_row=_BEA_CODE_ROW, values_only=True))  # type: ignore[attr-defined]
    for col_idx, code_cell in enumerate(code_row):
        if col_idx < _FIRST_INDUSTRY_COL - 1:
            continue
        if isinstance(code_cell, str) and code_cell.strip():
            industry_col_to_bea_code[col_idx] = code_cell.strip()

    # Second pass: snapshot the GO row for denominators. IOUse XLSX puts the
    # total-industry-output marker in the Name column when IOCode is None
    # (unlike Supply-Use which uses IOCode='T018' directly).
    gross_output_row: tuple[object, ...] | None = None
    for row in ws.iter_rows(min_row=8, values_only=True):  # type: ignore[attr-defined]
        iocode = row[_IOCODE_COL - 1] if len(row) >= _IOCODE_COL else None
        name = row[_NAME_COL - 1] if len(row) >= _NAME_COL else None
        if isinstance(iocode, str) and iocode.strip() == _STUB_GROSS_OUTPUT_IOCODE:
            gross_output_row = row
            break
        if iocode is None and isinstance(name, str) and name.strip() == _STUB_GROSS_OUTPUT_NAME:
            gross_output_row = row
            break

    if gross_output_row is None:
        raise BEAIngestError(
            f"year {year}: missing T018 / 'Total Industry Output' row in IOUse XLSX "
            "— cannot normalize coefficients"
        )

    go_by_col: dict[int, Decimal] = {}
    for col_idx in industry_col_to_bea_code:
        go = _coerce_cell_to_decimal(gross_output_row[col_idx])
        if go is not None and go > Decimal("0"):
            go_by_col[col_idx] = go

    # Third pass: iterate every data row, emit a coefficient record per non-zero cell
    for row in ws.iter_rows(min_row=8, values_only=True):  # type: ignore[attr-defined]
        iocode = row[_IOCODE_COL - 1] if len(row) >= _IOCODE_COL else None
        if not isinstance(iocode, str):
            continue
        iocode = iocode.strip()
        # Skip stub rows (T-prefixed totals, V-prefixed value-added, etc.)
        # Skip totals (T*-prefixed) and value-added (V*-prefixed) rows.
        # KEEP 'Used' and 'Other' — they are legitimate source industries
        # in dim_bea_industry (ids 106 and 107) and contribute to FR-004
        # column sums per BEA's accounting convention.
        if not iocode or iocode.startswith(("T", "V")):
            continue

        source_bea_id = bea_code_lookup.get(iocode)
        if source_bea_id is None:
            continue

        for col_idx, target_bea_code in industry_col_to_bea_code.items():
            target_bea_id = bea_code_lookup.get(target_bea_code)
            if target_bea_id is None:
                continue
            go = go_by_col.get(col_idx)
            if go is None:
                continue
            use_dollars = _coerce_cell_to_decimal(row[col_idx])
            if use_dollars is None or use_dollars == Decimal("0"):
                continue  # sparsity policy

            coefficient = float(use_dollars / go)
            if coefficient <= 0.0:
                continue
            # Cap at 1.5 per data-model.md §Entity 2 (admits recycled-input cases).
            if coefficient > 1.5:
                log.warning(
                    "year %d: coefficient out-of-bounds: src=%s tgt=%s value=%.6f — clamping to 1.5",
                    year,
                    iocode,
                    target_bea_code,
                    coefficient,
                )
                coefficient = 1.5

            yield BEAIOCoefficientRecord(
                source_industry_id=source_bea_id,
                target_industry_id=target_bea_id,
                table_type=table_type,  # type: ignore[arg-type]
                year=year,
                coefficient=coefficient,
                vintage_published_date=vintage,  # type: ignore[arg-type]
            )


def extract_iouse_internal_shares(  # noqa: C901 — XLSX multi-pass parser
    xlsx_path: Path,
    years: range,
    session: Session,
) -> dict[tuple[int, int], float]:
    """Read IOUse's own T005 (intermediate) / T018 (gross output) per (industry, year).

    Returns the producer-prices intermediate-inputs share for each
    target industry / year, used by ``validate_column_sum_identity``
    as the expected reference value for FR-004.

    Args:
        xlsx_path: Path to IOUse_Before_Redefinitions_PRO_Summary.xlsx.
        years: Year range.
        session: SQLAlchemy session for bea_code lookups.

    Returns:
        Mapping ``(target_industry_id, year) -> intermediate_inputs_share``.
    """
    if not xlsx_path.exists():
        raise BEAIngestError(f"BEA IOUse XLSX not found at {xlsx_path!r}")

    bea_rows = session.execute(
        select(DimBEAIndustry.bea_code, DimBEAIndustry.bea_industry_id)
    ).all()
    bea_code_lookup: dict[str, int] = {code: bid for code, bid in bea_rows}  # noqa: C416
    year_rows = session.execute(
        select(DimTime.year, DimTime.time_id).where(DimTime.is_annual.is_(True))
    ).all()
    year_lookup: dict[int, int] = {year: tid for year, tid in year_rows}  # noqa: C416
    years_in_scope = set(years)

    shares: dict[tuple[int, int], float] = {}

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                continue
            if year not in years_in_scope or year not in year_lookup:
                continue
            ws = wb[sheet_name]

            # Industry columns
            industry_col_to_bea_code: dict[int, str] = {}
            code_row = next(
                ws.iter_rows(min_row=_BEA_CODE_ROW, max_row=_BEA_CODE_ROW, values_only=True)
            )
            for col_idx, code_cell in enumerate(code_row):
                if col_idx < _FIRST_INDUSTRY_COL - 1:
                    continue
                if isinstance(code_cell, str) and code_cell.strip():
                    industry_col_to_bea_code[col_idx] = code_cell.strip()

            # Total Intermediate row + Total Industry Output row
            intermediate_row: tuple[object, ...] | None = None
            gross_output_row: tuple[object, ...] | None = None
            for row in ws.iter_rows(min_row=8, values_only=True):
                iocode = row[_IOCODE_COL - 1] if len(row) >= _IOCODE_COL else None
                name = row[_NAME_COL - 1] if len(row) >= _NAME_COL else None
                if iocode is None and isinstance(name, str):
                    n = name.strip()
                    if n == "Total Intermediate":
                        intermediate_row = row
                    elif n == _STUB_GROSS_OUTPUT_NAME:
                        gross_output_row = row
                if intermediate_row is not None and gross_output_row is not None:
                    break

            if intermediate_row is None or gross_output_row is None:
                log.warning(
                    "year %d: IOUse missing 'Total Intermediate' or 'Total Industry Output' "
                    "row — skipping internal shares extraction",
                    year,
                )
                continue

            for col_idx, bea_code in industry_col_to_bea_code.items():
                target_id = bea_code_lookup.get(bea_code)
                if target_id is None:
                    continue
                ii = _coerce_cell_to_decimal(intermediate_row[col_idx])
                go = _coerce_cell_to_decimal(gross_output_row[col_idx])
                if ii is None or go is None or go == Decimal("0"):
                    continue
                shares[(target_id, year)] = float(ii / go)
    finally:
        wb.close()

    return shares


def parse_total_req_matrix(  # noqa: C901 — XLSX multi-pass parser
    xlsx_path: Path,
    years: range,
    session: Session,
) -> Iterator[BEAIOCoefficientRecord]:
    """Parse Total Domestic Requirements (Leontief inverse) XLSX.

    Unlike Use_Summary, TDR cells are already coefficients (no GO
    normalization needed). TDR coefficients can exceed 1.0 (the
    diagonal entries are typically 1.x because they include the
    direct-output unit demand).

    Args:
        xlsx_path: Path to ``IxI_TR_Summary.xlsx``.
        years: Year range to ingest.
        session: SQLAlchemy session.

    Yields:
        ``BEAIOCoefficientRecord`` with ``table_type="TOTAL_REQ"``.
    """
    if not xlsx_path.exists():
        raise BEAIngestError(f"BEA TDR XLSX not found at {xlsx_path!r}")

    vintage = extract_vintage_date(xlsx_path)
    bea_rows = session.execute(
        select(DimBEAIndustry.bea_code, DimBEAIndustry.bea_industry_id)
    ).all()
    bea_code_lookup: dict[str, int] = {code: bid for code, bid in bea_rows}  # noqa: C416
    year_rows = session.execute(
        select(DimTime.year, DimTime.time_id).where(DimTime.is_annual.is_(True))
    ).all()
    year_lookup: dict[int, int] = {year: tid for year, tid in year_rows}  # noqa: C416
    years_in_scope = set(years)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                continue
            if year not in years_in_scope or year not in year_lookup:
                continue
            ws = wb[sheet_name]

            # Industry column lookup
            industry_col_to_bea_code: dict[int, str] = {}
            code_row = next(
                ws.iter_rows(min_row=_BEA_CODE_ROW, max_row=_BEA_CODE_ROW, values_only=True)
            )
            for col_idx, code_cell in enumerate(code_row):
                if col_idx < _FIRST_INDUSTRY_COL - 1:
                    continue
                if isinstance(code_cell, str) and code_cell.strip():
                    industry_col_to_bea_code[col_idx] = code_cell.strip()

            # Data rows
            for row in ws.iter_rows(min_row=8, values_only=True):
                iocode = row[_IOCODE_COL - 1] if len(row) >= _IOCODE_COL else None
                if not isinstance(iocode, str):
                    continue
                iocode = iocode.strip()
                if not iocode or iocode.startswith(("T", "V")):
                    continue
                source_bea_id = bea_code_lookup.get(iocode)
                if source_bea_id is None:
                    continue

                for col_idx, target_bea_code in industry_col_to_bea_code.items():
                    target_bea_id = bea_code_lookup.get(target_bea_code)
                    if target_bea_id is None:
                        continue
                    cell = row[col_idx]
                    coefficient = _coerce_cell_to_decimal(cell)
                    if coefficient is None or coefficient == Decimal("0"):
                        continue
                    # TDR diagonals can be > 1; cap at 1.5 for storage safety
                    coef_float = float(coefficient)
                    if coef_float > 1.5:
                        coef_float = 1.5
                    if coef_float <= 0.0:
                        continue

                    yield BEAIOCoefficientRecord(
                        source_industry_id=source_bea_id,
                        target_industry_id=target_bea_id,
                        table_type="TOTAL_REQ",
                        year=year,
                        coefficient=coef_float,
                        vintage_published_date=vintage,
                    )
    finally:
        wb.close()
