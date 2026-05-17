"""Parser for the BEA Supply-Use Use_Summary XLSX (spec-068 US1).

File path: ``data/input-output/supply-use/Use_Summary.xlsx``.
Sheet layout (per BEA convention, documented in ``data/bea/io_loader.py``):

* One sheet per year, named like ``'2010'``, ``'2011'``, ..., ``'2024'``.
* Header rows:

    * R1: Title string (e.g., "The Use of Commodities by Industries - Summary").
    * R2: Units string ``"(Millions of dollars)"`` (FR-002 sanity check).
    * R3: ``"Bureau of Economic Analysis"``.
    * R4: Year string (cross-check vs sheet name).
    * R6: Industry BEA codes in cols 3+ (``111CA``, ``113FF``, ``211``, ...).
    * R7: Industry names in cols 3+ (matches dim_bea_industry.industry_name).

* Stub rows (rightmost ``IOCode`` column == col 0):

    * ``T005``: Total Intermediate (``II``) — row's value for each
      industry column is the per-industry intermediate-inputs total.
    * ``VABAS``: Value Added (basic prices) (``VA``) — the
      Marxian-decomposition value-added is the basic-prices figure.
    * ``T018``: Total industry output (basic prices) (``GO``) —
      producer-side industry output per Clarification Q4.

The parser yields one ``BEAIndustryAnnualRecord`` per (industry, year)
cell that resolves to a valid bea_industry_id in ``dim_bea_industry``.
"""

from __future__ import annotations

import logging
from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.orm import Session

from babylon.reference.bea.ingest.vintage_extractor import extract_vintage_date
from babylon.reference.bea.models import BEAIndustryAnnualRecord
from babylon.reference.schema import DimBEAIndustry, DimTime

log = logging.getLogger(__name__)

_BEA_CODE_ROW = 6
_INDUSTRY_NAME_ROW = 7
_FIRST_INDUSTRY_COL = 3  # cols 1+2 are stub labels; col 3+ are industries
_IOCODE_COL = 1

_STUB_TOTAL_INTERMEDIATE = "T005"
_STUB_VALUE_ADDED = "VABAS"  # Value Added at basic prices (Marxian-aligned)
_STUB_GROSS_OUTPUT = "T018"  # Total industry output at basic prices

_REQUIRED_STUBS = frozenset({_STUB_TOTAL_INTERMEDIATE, _STUB_VALUE_ADDED, _STUB_GROSS_OUTPUT})


class BEAIngestError(RuntimeError):
    """Raised when a BEA XLSX cannot be parsed in a recoverable way."""


def _build_bea_code_lookup(session: Session) -> dict[str, int]:
    """Map ``bea_code`` (XLSX-format like ``111CA``) -> ``bea_industry_id``."""
    rows = session.execute(select(DimBEAIndustry.bea_code, DimBEAIndustry.bea_industry_id)).all()
    return {code: bid for code, bid in rows}  # noqa: C416  # Row[tuple] not tuple per mypy


def _build_year_lookup(session: Session) -> dict[int, int]:
    """Map ``year`` -> ``time_id`` for annual rows in ``dim_time``."""
    rows = session.execute(
        select(DimTime.year, DimTime.time_id).where(DimTime.is_annual.is_(True))
    ).all()
    return {year: tid for year, tid in rows}  # noqa: C416  # Row[tuple] not tuple per mypy


def _coerce_cell_to_decimal(value: object) -> Decimal | None:
    """Convert a BEA XLSX cell value to ``Decimal | None``.

    BEA marks suppressions with the string ``'...'`` — those map to ``None``.
    Numeric cells coerce via ``str(...)`` to preserve precision.
    """
    if value is None or value == "...":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped == "...":
            return None
        try:
            return Decimal(stripped)
        except Exception:  # noqa: BLE001
            return None
    return None


def parse_use_summary(
    xlsx_path: Path,
    years: range,
    session: Session,
) -> Iterator[BEAIndustryAnnualRecord]:
    """Yield one ``BEAIndustryAnnualRecord`` per (industry, year) in ``years``.

    Args:
        xlsx_path: Path to ``Use_Summary.xlsx`` (typically
            ``data/input-output/supply-use/Use_Summary.xlsx``).
        years: Iterable of years to ingest. Sheets outside this range
            are skipped.
        session: SQLAlchemy session for dim_bea_industry + dim_time lookups.

    Yields:
        Frozen Pydantic records ready for the UPSERT writer.

    Raises:
        BEAIngestError: If the XLSX is missing or unreadable, or if any
            year's sheet lacks the required stub rows (T005/VABAS/T018).
    """
    if not xlsx_path.exists():
        raise BEAIngestError(f"BEA Use_Summary XLSX not found at {xlsx_path!r}")

    vintage = extract_vintage_date(xlsx_path)
    bea_code_lookup = _build_bea_code_lookup(session)
    year_lookup = _build_year_lookup(session)
    years_in_scope = set(years)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                log.debug("skip non-year sheet %r", sheet_name)
                continue
            if year not in years_in_scope:
                continue
            if year not in year_lookup:
                log.warning("year %d not in dim_time — skipping (no time_id mapping)", year)
                continue

            ws = wb[sheet_name]
            yield from _parse_year_sheet(
                ws=ws,
                year=year,
                bea_code_lookup=bea_code_lookup,
                vintage=vintage,
            )
    finally:
        wb.close()


def _parse_year_sheet(
    ws: object,  # openpyxl Worksheet; openpyxl lacks public type
    year: int,
    bea_code_lookup: dict[str, int],
    vintage: object,
) -> Iterator[BEAIndustryAnnualRecord]:
    """Yield records for one year sheet of Use_Summary."""
    # First pass: discover industry column positions from header rows.
    industry_col_to_bea_code: dict[int, str] = {}
    code_row = next(ws.iter_rows(min_row=_BEA_CODE_ROW, max_row=_BEA_CODE_ROW, values_only=True))  # type: ignore[attr-defined]
    for col_idx, code_cell in enumerate(code_row):
        if col_idx < _FIRST_INDUSTRY_COL - 1:  # iter_rows is 0-indexed
            continue
        if not isinstance(code_cell, str) or not code_cell.strip():
            continue
        industry_col_to_bea_code[col_idx] = code_cell.strip()

    # Second pass: scan rows, find the three stub rows, snapshot their values.
    stub_rows: dict[str, tuple[object, ...]] = {}
    for row in ws.iter_rows(min_row=8, values_only=True):  # type: ignore[attr-defined]
        iocode = row[_IOCODE_COL - 1] if len(row) >= _IOCODE_COL else None
        if isinstance(iocode, str) and iocode.strip() in _REQUIRED_STUBS:
            stub_rows[iocode.strip()] = row

    missing = _REQUIRED_STUBS - stub_rows.keys()
    if missing:
        raise BEAIngestError(
            f"year {year}: missing required stub rows {sorted(missing)!r} "
            f"(expected T005, VABAS, T018)"
        )

    intermediate_row = stub_rows[_STUB_TOTAL_INTERMEDIATE]
    value_added_row = stub_rows[_STUB_VALUE_ADDED]
    gross_output_row = stub_rows[_STUB_GROSS_OUTPUT]

    for col_idx, bea_code in industry_col_to_bea_code.items():
        bea_industry_id = bea_code_lookup.get(bea_code)
        if bea_industry_id is None:
            log.debug(
                "year %d: unknown bea_code %r in column %d — skipping",
                year,
                bea_code,
                col_idx,
            )
            continue

        ii = _coerce_cell_to_decimal(intermediate_row[col_idx])
        va = _coerce_cell_to_decimal(value_added_row[col_idx])
        go = _coerce_cell_to_decimal(gross_output_row[col_idx])

        yield BEAIndustryAnnualRecord(
            bea_industry_id=bea_industry_id,
            year=year,
            gross_output_millions=go,
            intermediate_inputs_millions=ii,
            value_added_millions=va,
            vintage_published_date=vintage,  # type: ignore[arg-type]
        )
