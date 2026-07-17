"""Behavioral contract for the fact_productivity_annual loader (ADR075 ruling 1).

Owner ruling 1 flipped A11 fact_productivity_annual from AMPUTATE to FILL:
"keep fact_productivity_annual and try to fill it in." No loader has ever
existed for this table (the census verified that); this one reads the staged
BLS detailed-industries workbook
(``/media/user/data/babylon-data/productivity/labor-productivity-detailed-industries.xlsx``,
``MachineReadable`` sheet — long format: Sector/NAICS/Industry/Digit/Basis/
Measure/Units/Year/Value) and pivots ten (Measure, Units) combinations onto
the ``FactProductivityAnnual`` columns, joining NAICS -> dim_industry and
Year -> dim_time. Filling it revives ``view_surplus_value`` (rate of
exploitation s/v) and ``view_imperial_rent`` (W_c - V_c) — the Fundamental
Theorem's empirical legs.

All tests use synthetic workbooks + tmp sqlite fixtures — never the real DB
or the trove (CI never touches the babylon-data drive).
"""

from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

TOOLS_DIR = Path(__file__).resolve().parents[3] / "tools"
sys.path.insert(0, str(TOOLS_DIR))

from load_productivity_annual import (  # type: ignore[import-not-found]  # noqa: E402
    COLUMN_BY_MEASURE_UNITS,
    LoaderError,
    load_workbook_rows,
    main,
    pivot_rows,
)

_HEADER = ("Sector", "NAICS", "Industry", "Digit", "Basis", "Measure", "Units", "Year", "Value")


def _mini_workbook(path: Path) -> None:
    """Two NAICS codes x two years; one unknown NAICS; one N.A. value."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "MachineReadable"
    ws.append(_HEADER)
    rows = [
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2022,
            101.5,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Sectoral output",
            "Millions of current dollars",
            2022,
            500000.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Labor compensation",
            "Millions of current dollars",
            2022,
            120000.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Hours worked",
            "Millions of hours",
            2022,
            1300.25,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Employment",
            "Thousands of jobs",
            2022,
            600.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Hours worked",
            "Index (2017=100)",
            2022,
            98.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Hourly compensation",
            "Index (2017=100)",
            2022,
            110.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Unit labor costs",
            "Index (2017=100)",
            2022,
            104.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Real sectoral output",
            "Index (2017=100)",
            2022,
            103.0,
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Sectoral output",
            "Index (2017=100)",
            2022,
            105.0,
        ),
        # A second year with an N.A. value: stored as NULL, row still lands.
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2023,
            "N.A.",
        ),
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Labor compensation",
            "Millions of current dollars",
            2023,
            125000.0,
        ),
        # A measure the loader does not consume — silently out of scope.
        (
            "Mining",
            "21",
            "Mining",
            "2-Digit",
            "All workers",
            "Capital share",
            "Percentage",
            2022,
            40.0,
        ),
        # NAICS absent from dim_industry — skipped, reported.
        (
            "Utilities",
            "2211X",
            "Fake",
            "5-Digit",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2022,
            90.0,
        ),
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def _fixture_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        CREATE TABLE dim_industry (
            industry_id INTEGER PRIMARY KEY,
            naics_code VARCHAR(10) NOT NULL
        );
        INSERT INTO dim_industry VALUES (7, '21'), (8, '22');

        CREATE TABLE dim_time (
            time_id INTEGER PRIMARY KEY,
            year INTEGER NOT NULL,
            is_annual BOOLEAN NOT NULL
        );
        INSERT INTO dim_time VALUES (2022, 2022, 1), (2023, 2023, 1);

        CREATE TABLE fact_productivity_annual (
            industry_id INTEGER NOT NULL,
            time_id INTEGER NOT NULL,
            labor_productivity_index NUMERIC(10, 4),
            hours_worked_index NUMERIC(10, 4),
            hourly_compensation_index NUMERIC(10, 4),
            unit_labor_costs_index NUMERIC(10, 4),
            real_output_index NUMERIC(10, 4),
            sectoral_output_index NUMERIC(10, 4),
            hours_worked_millions NUMERIC(12, 2),
            employment_thousands NUMERIC(10, 2),
            labor_compensation_millions_usd NUMERIC(15, 2),
            sectoral_output_millions_usd NUMERIC(15, 2),
            PRIMARY KEY (industry_id, time_id)
        );
        """
    )
    conn.commit()
    conn.close()


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "mini-productivity.xlsx"
    _mini_workbook(path)
    return path


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    path = tmp_path / "mini.sqlite"
    _fixture_db(path)
    return path


class TestMapping:
    def test_ten_column_mapping_is_pinned(self) -> None:
        assert len(COLUMN_BY_MEASURE_UNITS) == 10
        assert COLUMN_BY_MEASURE_UNITS[("Labor compensation", "Millions of current dollars")] == (
            "labor_compensation_millions_usd"
        )
        assert COLUMN_BY_MEASURE_UNITS[("Sectoral output", "Millions of current dollars")] == (
            "sectoral_output_millions_usd"
        )
        assert COLUMN_BY_MEASURE_UNITS[("Hours worked", "Index (2017=100)")] == "hours_worked_index"

    def test_pivot_collects_all_mapped_values(self, workbook_path: Path) -> None:
        rows = load_workbook_rows(workbook_path)
        pivot = pivot_rows(rows)
        cell = pivot[("21", 2022)]
        assert cell["labor_productivity_index"] == 101.5
        assert cell["sectoral_output_millions_usd"] == 500000.0
        assert cell["labor_compensation_millions_usd"] == 120000.0
        assert cell["hours_worked_millions"] == 1300.25
        assert cell["employment_thousands"] == 600.0
        assert len(pivot[("21", 2022)]) == 10

    def test_na_values_become_none(self, workbook_path: Path) -> None:
        pivot = pivot_rows(load_workbook_rows(workbook_path))
        cell = pivot[("21", 2023)]
        assert cell["labor_productivity_index"] is None
        assert cell["labor_compensation_millions_usd"] == 125000.0


class TestMainCLI:
    def test_dry_run_is_default_and_writes_nothing(
        self, workbook_path: Path, db_path: Path
    ) -> None:
        exit_code = main(["--source", str(workbook_path), "--db", str(db_path)])
        assert exit_code == 0
        conn = sqlite3.connect(db_path)
        count = conn.execute("SELECT COUNT(*) FROM fact_productivity_annual").fetchone()[0]
        conn.close()
        assert count == 0

    def test_execute_fills_exactly_the_joinable_pairs(
        self, workbook_path: Path, db_path: Path
    ) -> None:
        exit_code = main(["--source", str(workbook_path), "--db", str(db_path), "--execute"])
        assert exit_code == 0
        conn = sqlite3.connect(db_path)
        rows = conn.execute(
            "SELECT industry_id, time_id, labor_productivity_index,"
            " labor_compensation_millions_usd, sectoral_output_millions_usd"
            " FROM fact_productivity_annual ORDER BY time_id"
        ).fetchall()
        conn.close()
        assert len(rows) == 2  # NAICS 21 x {2022, 2023}; 2211X skipped
        assert rows[0] == (7, 2022, 101.5, 120000.0, 500000.0)
        assert rows[1] == (7, 2023, None, 125000.0, None)

    def test_nonempty_table_aborts_loudly(self, workbook_path: Path, db_path: Path) -> None:
        conn = sqlite3.connect(db_path)
        conn.execute("INSERT INTO fact_productivity_annual (industry_id, time_id) VALUES (8, 2022)")
        conn.commit()
        conn.close()
        with pytest.raises(LoaderError, match="not empty"):
            main(["--source", str(workbook_path), "--db", str(db_path), "--execute"])

    def test_missing_year_in_dim_time_aborts_loudly(
        self, workbook_path: Path, db_path: Path
    ) -> None:
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM dim_time WHERE year = 2023")
        conn.commit()
        conn.close()
        with pytest.raises(LoaderError, match="2023"):
            main(["--source", str(workbook_path), "--db", str(db_path), "--execute"])

    def test_missing_source_aborts_loudly(self, db_path: Path, tmp_path: Path) -> None:
        with pytest.raises(LoaderError, match="not found"):
            main(["--source", str(tmp_path / "nope.xlsx"), "--db", str(db_path), "--execute"])
