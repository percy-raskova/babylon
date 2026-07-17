#!/usr/bin/env python3
"""Extract the owner-approved empirical-invariant series into in-repo CSVs.

Owner ruling 2026-07-16: three empirical laws are approved as test-encoded
invariants — Piketty's β (wealth/income ratio), the US aggregate labor share,
and the imperial-rent multiple. Their raw sources live on the babylon-data
drive (WID country dump; BEA Use_Summary workbook), which CI must never
touch, so this extractor derives small per-year series and writes them to
``src/babylon/data/reference/`` where the contract tests
(``tests/unit/reference/test_empirical_series_artifacts.py``) read them.

Run locally whenever the upstream data revs::

    poetry run python -m tools.extract.empirical_invariant_series

Outputs (committed artifacts):

* ``wid_us_beta.csv`` — year, beta (US net national wealth / national income,
  WID ``wnweali999``), 1800–2024.
* ``wid_imperial_rent_multiple.csv`` — year, US mean per-adult income, US
  bottom-50% average income (``aptincj992`` p0p50), Sub-Saharan-Africa mean
  per-adult income (XF-MER aggregate file — WID's own USD series; single-
  country LCU/FX division is a documented 17× trap), and the two ratios.
* ``bea_us_labor_share.csv`` — year, compensation (V001), value added
  (VABAS), labor_share = V001/VABAS summed over industry columns, 1997–2024.
"""

from __future__ import annotations

import csv
import sys
from collections.abc import Sequence
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = _REPO_ROOT / "src" / "babylon" / "data" / "reference"

WID_DIR = Path("/media/user/data/babylon-data/piketty")
BEA_USE_XLSX = Path("/media/user/data/babylon-data/input-output/supply-use/Use_Summary.xlsx")


def _wid_series(country_file: Path, variable: str, percentile: str) -> dict[int, float]:
    """Read one (variable, percentile) annual series from a WID country CSV."""
    series: dict[int, float] = {}
    with country_file.open() as fh:
        for row in csv.reader(fh, delimiter=";"):
            if len(row) >= 5 and row[1] == variable and row[2] == percentile:
                try:
                    series[int(row[3])] = float(row[4])
                except ValueError:
                    continue
    if not series:
        msg = f"no rows for {variable}/{percentile} in {country_file}"
        raise RuntimeError(msg)
    return series


def extract_beta(out_path: Path) -> int:
    beta = _wid_series(WID_DIR / "WID_data_US.csv", "wnweali999", "p0p100")
    with out_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["year", "beta"])
        for year in sorted(beta):
            writer.writerow([year, f"{beta[year]:.6f}"])
    return len(beta)


def extract_imperial_multiple(out_path: Path) -> int:
    us = WID_DIR / "WID_data_US.csv"
    ssa = WID_DIR / "WID_data_XF-MER.csv"
    us_income = _wid_series(us, "mnninci999", "p0p100")
    us_adults = _wid_series(us, "npopuli992", "p0p100")
    us_bottom50 = _wid_series(us, "aptincj992", "p0p50")
    ssa_income = _wid_series(ssa, "mnninci999", "p0p100")
    ssa_adults = _wid_series(ssa, "npopuli992", "p0p100")

    years = sorted(
        set(us_income) & set(us_adults) & set(us_bottom50) & set(ssa_income) & set(ssa_adults)
    )
    years = [y for y in years if y >= 1950]
    if not years:
        raise RuntimeError("no overlapping years for the imperial-rent multiple")

    with out_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "year",
                "us_mean_income_per_adult_usd",
                "us_bottom50_avg_income_usd",
                "ssa_mean_income_per_adult_usd",
                "multiple_us_mean_over_ssa_mean",
                "multiple_us_bottom50_over_ssa_mean",
            ]
        )
        for year in years:
            us_pa = us_income[year] / us_adults[year]
            ssa_pa = ssa_income[year] / ssa_adults[year]
            writer.writerow(
                [
                    year,
                    f"{us_pa:.2f}",
                    f"{us_bottom50[year]:.2f}",
                    f"{ssa_pa:.2f}",
                    f"{us_pa / ssa_pa:.4f}",
                    f"{us_bottom50[year] / ssa_pa:.4f}",
                ]
            )
    return len(years)


def extract_labor_share(out_path: Path) -> int:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    if not BEA_USE_XLSX.exists():
        raise RuntimeError(f"BEA Use_Summary workbook not found at {BEA_USE_XLSX}")
    workbook = load_workbook(BEA_USE_XLSX, read_only=True, data_only=True)

    rows_out: list[tuple[int, float, float]] = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.strip().isdigit():
            continue
        year = int(sheet_name.strip())
        sheet = workbook[sheet_name]
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
        # Row 6 (1-based) carries industry BEA codes from col 3 (1-based) on;
        # sum only industry columns (skip totals T### and final-demand F###).
        code_row = grid[5]
        industry_cols: list[int] = []
        for idx in range(2, len(code_row)):
            code = code_row[idx]
            if isinstance(code, str) and code.strip() and not code.strip().startswith(("T", "F")):
                industry_cols.append(idx)

        def _stub_sum(stub: str, current: Sequence[Sequence[object]], cols: list[int]) -> float:
            for row in current:
                labels = {str(c).strip() for c in row[:3] if c is not None}
                if stub in labels:
                    total = 0.0
                    for i in cols:
                        cell = row[i]
                        if isinstance(cell, (int, float)):
                            total += float(cell)
                    return total
            raise RuntimeError(f"stub row {stub} not found in sheet")

        compensation = _stub_sum("V001", grid, industry_cols)
        value_added = _stub_sum("VABAS", grid, industry_cols)
        rows_out.append((year, compensation, value_added))

    if not rows_out:
        raise RuntimeError("no year sheets found in the BEA workbook")

    rows_out.sort()
    with out_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            ["year", "compensation_millions_usd", "value_added_millions_usd", "labor_share"]
        )
        for year, comp, va in rows_out:
            writer.writerow([year, f"{comp:.0f}", f"{va:.0f}", f"{comp / va:.6f}"])
    return len(rows_out)


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    n_beta = extract_beta(OUTPUT_DIR / "wid_us_beta.csv")
    print(f"wid_us_beta.csv: {n_beta} years")
    n_mult = extract_imperial_multiple(OUTPUT_DIR / "wid_imperial_rent_multiple.csv")
    print(f"wid_imperial_rent_multiple.csv: {n_mult} years")
    n_ls = extract_labor_share(OUTPUT_DIR / "bea_us_labor_share.csv")
    print(f"bea_us_labor_share.csv: {n_ls} years")
    return 0


if __name__ == "__main__":
    sys.exit(main())
