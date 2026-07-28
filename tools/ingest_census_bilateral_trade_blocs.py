#!/usr/bin/env python3
"""Ingest India / South-and-Central-America bilateral trade rows (Program 26 U3).

``fact_bilateral_trade_annual`` shipped with 120 rows (8 blocs x 2010-2024,
spec-100) but never included India (``dim_country.country_id=149``) or
South and Central America (``country_id=6``, the DB row backing the engine's
``latin_america`` node) — the ADR055-disclosed "india/latin_america -> Phi=0"
coverage hole. Both blocs are already present in ``dim_country``; only the
fact rows are missing ("schema-only since spec-100; no dedicated loader
script, populated additively" per ``data-catalog.yaml``). This is that
loader.

Source: ``/media/user/data/babylon-data/imperial_rent/country.xlsx`` — the
Census FT900 country/area table (the SAME source family that produced the
existing 120 rows; columns ``year, CTY_CODE, CTYNAME``, then 12 monthly +
1 annual (``IYR``) import columns, then 12 monthly + 1 annual (``EYR``)
export columns, in millions of current USD). ``total_trade_usd_millions =
IYR + EYR`` for the target ``year`` row, matching how the existing rows are
computed (verified: recomputing European Union/2015 from this file against
the DB's existing row reproduces it exactly to 6 significant figures — see
``tests/unit/persistence/test_phi_attribution.py``'s india/latin_america
conservation tests and the ceremony commit message for the full cross-check).

Rows targeted: ``CTY_CODE in ('0009', '5330')`` (South and Central America,
India), years 2010-2024 inclusive — the SAME 15-year annual window the
existing 8 blocs cover (``dim_time`` already has annual ``time_id`` rows for
every one of those years; no new ``dim_time`` rows are created here).
Idempotent: an ``(time_id, country_id)`` pair already present in the target
table is skipped, not duplicated (defensive — the ``loader_to_sources.py``
wrapper always runs against a fresh scratch copy, so this should never fire
in normal use, but a stray re-run must not violate the PK).

Usage (build-time only; requires the imperial_rent drive path)::

    uv run python tools/ingest_census_bilateral_trade_blocs.py \\
        --db-url "sqlite:////home/user/projects/game/babylon/data/sqlite/marxist-data-3NF.sqlite"

.. note::
   Post-cutover (ADR098), normal usage goes through
   ``tools/loader_to_sources.py --loader ingest_census_bilateral_trade_blocs
   --tables fact_bilateral_trade_annual`` instead of calling ``main`` here
   directly — the wrapper runs this module's ``main`` against a SCRATCH COPY
   of the build product, re-exports the affected table as a parquet source,
   and regenerates the manifest.
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from babylon.reference.schema import DimTime, FactBilateralTradeAnnual

DEFAULT_COUNTRY_XLSX = Path("/media/user/data/babylon-data/imperial_rent/country.xlsx")
DB_URL = "sqlite:///marxist-data-3NF.sqlite"

#: cty_code (workbook-padded form) -> country_id. Every target is already
#: seeded in dim_country; this loader only adds fact rows. P26 U5c
#: (ADR165 Q4 disjoint taxonomy + Q3 Mexico->latin_america) extends the
#: original two U3 targets with the DISJOINT partner set backing the 8
#: INTERNATIONAL_NODES: individual countries + genuinely non-overlapping
#: aggregates, so no dollar of US trade is counted under two nodes (the
#: containing-bloc taxonomy's 138.6%-of-world-trade denominator dies with
#: the U5d crosswalk that consumes these rows). Values must stay UNIQUE —
#: pinned by test_target_map_is_disjoint.
_TARGET_CTY_CODES: dict[str, int] = {
    # latin_america (S&C America excludes Mexico in the Census taxonomy;
    # Mexico joins it by ADR165 Q3 ruling)
    "0009": 6,  # South and Central America
    "2010": 21,  # Mexico
    # india
    "5330": 149,  # India
    # canada (shrinks off the North America aggregate, ADR165 Q3)
    "1220": 19,  # Canada
    # china (individual row — no more Asia/Pacific-Rim double-count)
    "5700": 168,  # China
    # sub_saharan_africa (disjoint Census aggregate, replaces "Africa")
    "0019": 15,  # Sub Saharan Africa
    # russia_csi: Russia + the 11 CIS/CSI states
    "4621": 96,  # Russia
    "4622": 97,  # Belarus
    "4623": 98,  # Ukraine
    "4631": 99,  # Armenia
    "4632": 100,  # Azerbaijan
    "4633": 101,  # Georgia
    "4634": 102,  # Kazakhstan
    "4635": 103,  # Kyrgyzstan
    "4641": 104,  # Moldova
    "4642": 105,  # Tajikistan
    "4643": 106,  # Turkmenistan
    "4644": 107,  # Uzbekistan
    # southeast_asia: the 10 ASEAN members (no ASEAN aggregate exists in
    # dim_country; Oceania deliberately NOT folded in — the FAF artifact's
    # zone-808 note stays a freight-side disclosure)
    "5460": 154,  # Burma
    "5490": 155,  # Thailand
    "5520": 156,  # Vietnam
    "5530": 157,  # Laos
    "5550": 158,  # Cambodia
    "5570": 159,  # Malaysia
    "5590": 160,  # Singapore
    "5600": 161,  # Indonesia
    "5610": 163,  # Brunei
    "5650": 164,  # Philippines
}
_YEAR_MIN, _YEAR_MAX = 2010, 2024

# country.xlsx column indices (0-indexed; header row verified: year,
# CTY_CODE, CTYNAME, IJAN..IDEC, IYR, EJAN..EDEC, EYR).
_COL_YEAR = 0
_COL_CTY_CODE = 1
_COL_IMPORTS_ANNUAL = 15  # IYR
_COL_EXPORTS_ANNUAL = 28  # EYR


class BilateralTradeIngestError(Exception):
    """A loader step failed loudly — malformed source, missing rows."""


def read_bloc_annual_rows(xlsx_path: Path) -> dict[tuple[str, int], tuple[float, float]]:
    """Read ``{(cty_code, year): (imports_usd_millions, exports_usd_millions)}``
    for the target blocs / year window from ``country.xlsx``.

    :raises BilateralTradeIngestError: If ``xlsx_path`` is missing or has no
        usable header.
    """
    if not xlsx_path.is_file():
        msg = f"country.xlsx not found at {xlsx_path}"
        raise BilateralTradeIngestError(msg)

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    out: dict[tuple[str, int], tuple[float, float]] = {}
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_index == 0:
            if row[_COL_CTY_CODE] != "CTY_CODE":
                msg = f"unexpected header in {xlsx_path}: {row!r}"
                raise BilateralTradeIngestError(msg)
            continue
        cty_code = row[_COL_CTY_CODE]
        if cty_code not in _TARGET_CTY_CODES:
            continue
        year_raw = row[_COL_YEAR]
        if year_raw is None or not str(year_raw).isdigit():
            continue
        year = int(year_raw)
        if not (_YEAR_MIN <= year <= _YEAR_MAX):
            continue
        imports = row[_COL_IMPORTS_ANNUAL]
        exports = row[_COL_EXPORTS_ANNUAL]
        if imports is None or exports is None:
            continue
        out[(cty_code, year)] = (float(imports), float(exports))
    workbook.close()
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db-url", default=DB_URL, help="Database URL")
    parser.add_argument("--country-xlsx", type=Path, default=DEFAULT_COUNTRY_XLSX)
    args = parser.parse_args(argv)

    try:
        bloc_rows = read_bloc_annual_rows(args.country_xlsx)
    except BilateralTradeIngestError as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    if not bloc_rows:
        print(f"Error: no target bloc/year rows found in {args.country_xlsx}", file=sys.stderr)
        return 1

    engine = create_engine(args.db_url)
    with Session(engine) as session:
        try:
            time_id_by_year: dict[int, int] = {}
            for year in range(_YEAR_MIN, _YEAR_MAX + 1):
                time_row = session.execute(
                    select(DimTime.time_id).where(DimTime.year == year, DimTime.is_annual == True)  # noqa: E712
                ).scalar_one_or_none()
                if time_row is None:
                    print(f"Error: no annual dim_time row for year={year}", file=sys.stderr)
                    return 1
                time_id_by_year[year] = time_row

            existing = {
                (time_id, country_id)
                for time_id, country_id in session.execute(
                    select(
                        FactBilateralTradeAnnual.time_id, FactBilateralTradeAnnual.country_id
                    ).where(FactBilateralTradeAnnual.country_id.in_(_TARGET_CTY_CODES.values()))
                ).all()
            }

            inserted = 0
            for (cty_code, year), (imports, exports) in sorted(bloc_rows.items()):
                country_id = _TARGET_CTY_CODES[cty_code]
                time_id = time_id_by_year[year]
                if (time_id, country_id) in existing:
                    continue
                session.add(
                    FactBilateralTradeAnnual(
                        time_id=time_id,
                        country_id=country_id,
                        imports_usd_millions=Decimal(str(round(imports, 2))),
                        exports_usd_millions=Decimal(str(round(exports, 2))),
                        total_trade_usd_millions=Decimal(str(round(imports + exports, 2))),
                    )
                )
                inserted += 1
            session.commit()
            print(f"Inserted {inserted} fact_bilateral_trade_annual rows (bilateral partners).")
            # Loud per-code coverage report (U5c): a target code with zero
            # in-window rows is disclosed, never silently absent — the
            # ceremony operator verifies this table against expectations.
            per_code = dict.fromkeys(_TARGET_CTY_CODES, 0)
            for cty_code, _year in bloc_rows:
                per_code[cty_code] += 1
            for code in sorted(per_code):
                marker = "" if per_code[code] else "  <-- NO IN-WINDOW ROWS"
                print(f"  cty_code {code}: {per_code[code]} year-rows{marker}")
        except Exception as error:  # noqa: BLE001 - loader boundary: report and roll back loudly
            session.rollback()
            print(f"Error ingesting bilateral trade rows: {error}", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
