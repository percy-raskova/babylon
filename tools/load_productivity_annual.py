"""One-shot fill of ``fact_productivity_annual`` from the staged BLS workbook.

ADR075 ruling 1 (owner, 2026-07-16) flipped this table from AMPUTATE to FILL:
"keep fact_productivity_annual and try to fill it in." No loader had ever
targeted it — ``view_surplus_value`` (rate of exploitation s/v) and
``view_imperial_rent`` (W_c − V_c, the Fundamental Theorem's gap) selected
from a permanently-empty base, the pathology that seeded the whole Data
Constitution program.

Source: the BLS "labor productivity and costs, detailed industries" workbook
staged in the trove (``MachineReadable`` sheet, long format). Ten
(Measure, Units) combinations pivot onto the ``FactProductivityAnnual``
columns; NAICS joins ``dim_industry.naics_code``; Year joins annual
``dim_time`` rows. NAICS codes absent from ``dim_industry`` are skipped and
reported (honest absence); a Year absent from ``dim_time`` aborts loudly —
the time dimension is shared state this tool must not mutate.

Usage::

    poetry run python tools/load_productivity_annual.py             # dry run
    poetry run python tools/load_productivity_annual.py --execute
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

#: The ten consumed (Measure, Units) combinations -> ORM column names.
#: Pinned by ``tests/unit/tools/test_load_productivity_annual.py``.
COLUMN_BY_MEASURE_UNITS: dict[tuple[str, str], str] = {
    ("Labor productivity", "Index (2017=100)"): "labor_productivity_index",
    ("Hours worked", "Index (2017=100)"): "hours_worked_index",
    ("Hourly compensation", "Index (2017=100)"): "hourly_compensation_index",
    ("Unit labor costs", "Index (2017=100)"): "unit_labor_costs_index",
    ("Real sectoral output", "Index (2017=100)"): "real_output_index",
    ("Sectoral output", "Index (2017=100)"): "sectoral_output_index",
    ("Hours worked", "Millions of hours"): "hours_worked_millions",
    ("Employment", "Thousands of jobs"): "employment_thousands",
    ("Labor compensation", "Millions of current dollars"): "labor_compensation_millions_usd",
    ("Sectoral output", "Millions of current dollars"): "sectoral_output_millions_usd",
}

_COLUMNS: tuple[str, ...] = tuple(COLUMN_BY_MEASURE_UNITS.values())

_DEFAULT_SOURCE = Path(
    "/media/user/data/babylon-data/productivity/labor-productivity-detailed-industries.xlsx"
)
_DEFAULT_DB = Path("data/sqlite/marxist-data-3NF.sqlite")
_ENV_OVERRIDE = "BABYLON_NORMALIZED_DB_PATH"


class LoaderError(Exception):
    """A pre-flight check failed; the database was not modified."""


def load_workbook_rows(source: Path) -> list[tuple[str, int, str, str, float | None]]:
    """Read the ``MachineReadable`` sheet into (naics, year, measure, units, value) rows.

    :param source: Path to the staged BLS workbook.
    :returns: One tuple per consumed sheet row; ``N.A.``/blank values are ``None``.
    :raises LoaderError: When the file or sheet is missing.
    """
    if not source.exists():
        msg = f"source workbook not found: {source}"
        raise LoaderError(msg)
    workbook = load_workbook(source, read_only=True)
    if "MachineReadable" not in workbook.sheetnames:
        msg = f"sheet 'MachineReadable' missing from {source}"
        raise LoaderError(msg)
    sheet = workbook["MachineReadable"]
    rows: list[tuple[str, int, str, str, float | None]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        naics, measure, units, year, value = raw[1], raw[5], raw[6], raw[7], raw[8]
        if naics is None or year is None:
            continue
        rows.append((str(naics), int(str(year)), str(measure), str(units), _coerce(value)))
    return rows


def _coerce(value: object) -> float | None:
    """Parse a sheet cell: numeric, numeric-string, or the BLS ``N.A.`` marker."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip()
    if text in ("", "N.A.", "NA", "-"):
        return None
    return float(text)


def pivot_rows(
    rows: list[tuple[str, int, str, str, float | None]],
) -> dict[tuple[str, int], dict[str, float | None]]:
    """Pivot long-format rows onto the fact-table column set.

    :param rows: Output of :func:`load_workbook_rows`.
    :returns: ``{(naics, year): {column: value}}`` with every consumed column
        present (missing combinations stay ``None``).
    """
    pivot: dict[tuple[str, int], dict[str, float | None]] = {}
    for naics, year, measure, units, value in rows:
        column = COLUMN_BY_MEASURE_UNITS.get((measure, units))
        if column is None:
            continue  # 17 of the 27 BLS measures are out of scope by design
        cell = pivot.setdefault((naics, year), dict.fromkeys(_COLUMNS))
        cell[column] = value
    return pivot


def _resolve_dimensions(
    conn: sqlite3.Connection, pivot: dict[tuple[str, int], dict[str, float | None]]
) -> tuple[dict[str, int], dict[int, int], list[str]]:
    """Map NAICS -> industry_id and year -> time_id for the pivoted pairs.

    :returns: ``(industry_by_naics, time_by_year, skipped_naics)``.
    :raises LoaderError: When any year lacks an annual ``dim_time`` row.
    """
    industry_by_naics = {
        str(code): int(industry_id)
        for industry_id, code in conn.execute("SELECT industry_id, naics_code FROM dim_industry")
    }
    time_by_year = {
        int(year): int(time_id)
        for time_id, year in conn.execute("SELECT time_id, year FROM dim_time WHERE is_annual = 1")
    }
    needed_years = {year for _, year in pivot}
    missing_years = sorted(needed_years - set(time_by_year))
    if missing_years:
        msg = f"dim_time lacks annual rows for years {missing_years} — refusing to mutate dims"
        raise LoaderError(msg)
    skipped = sorted({naics for naics, _ in pivot if naics not in industry_by_naics})
    return industry_by_naics, time_by_year, skipped


def main(argv: list[str] | None = None) -> int:
    """Run the fill (dry run unless ``--execute``).

    :param argv: CLI arguments (defaults to ``sys.argv[1:]``).
    :returns: ``0`` on success.
    :raises LoaderError: When a pre-flight check fails; nothing is written.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=_DEFAULT_SOURCE)
    parser.add_argument(
        "--db",
        type=Path,
        default=Path(os.environ.get(_ENV_OVERRIDE, str(_DEFAULT_DB))),
        help="Path to the reference DB (default honors BABYLON_NORMALIZED_DB_PATH).",
    )
    parser.add_argument(
        "--execute", action="store_true", help="Actually insert (default: dry run)."
    )
    args = parser.parse_args(argv)

    if not args.db.exists():
        msg = f"database not found: {args.db}"
        raise LoaderError(msg)
    pivot = pivot_rows(load_workbook_rows(args.source))

    conn = sqlite3.connect(args.db)
    try:
        existing = int(conn.execute("SELECT COUNT(*) FROM fact_productivity_annual").fetchone()[0])
        if existing:
            msg = f"fact_productivity_annual is not empty ({existing} rows) — one-shot fill refused"
            raise LoaderError(msg)
        industry_by_naics, time_by_year, skipped = _resolve_dimensions(conn, pivot)
        insertable = [
            (industry_by_naics[naics], time_by_year[year], *(cell[c] for c in _COLUMNS))
            for (naics, year), cell in sorted(pivot.items())
            if naics in industry_by_naics
        ]
        print(
            f"[productivity] {len(insertable)} (industry, year) rows ready; "
            f"{len(skipped)} NAICS codes absent from dim_industry: {skipped}"
        )
        if not args.execute:
            print("[productivity] dry run — nothing written (pass --execute to proceed)")
            return 0
        placeholders = ", ".join(["?"] * (2 + len(_COLUMNS)))
        columns = ", ".join(("industry_id", "time_id", *_COLUMNS))
        conn.executemany(
            f"INSERT INTO fact_productivity_annual ({columns}) VALUES ({placeholders})",  # noqa: S608
            insertable,
        )
        conn.commit()
        count = int(conn.execute("SELECT COUNT(*) FROM fact_productivity_annual").fetchone()[0])
        print(f"[productivity] inserted {count} rows")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except LoaderError as error:
        print(f"[productivity] ABORT: {error}", file=sys.stderr)
        sys.exit(2)
